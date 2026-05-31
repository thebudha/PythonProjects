import openpyxl

#wb = openpyxl.load_workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["A1"]
print(sheet.max_row)
print(sheet.max_column)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=column)
        print(cell.value)